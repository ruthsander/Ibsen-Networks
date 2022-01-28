import xml.etree.ElementTree as ET
from lxml import etree
from html.entities import codepoint2name
import pprint
import re
import csv
from pywikibot.data import api
import pywikibot
import requests
from pywikibot import pagegenerators as pg
import pandas as pd
from csv import writer
from csv import reader
import time
import numpy as np
import sys
import openpyxl

#  declaring all namespaces present in the xml files
ns = {'HIS': 'http://www.example.org/ns/HIS',
      'xml': 'http://www.w3.org/XML/1998/namespace',
      'tei': 'http://www.tei-c.org/ns/1.0'}

tree = etree.parse('../github/Ibsen-Networks/xml-data/Navneregister_HISe.xml')
regest_entries = tree.xpath('//tei:div[@xml:id and @type="person" or @type="organisation"]', namespaces=ns)
persons_regest_entries = tree.xpath('//tei:div[@xml:id and @type="person"]', namespaces=ns)

# colnames = ['XML_ID','Wikidata_ID','Viaf_ID','NHRP_ID','Given_Name','Surname','Name','Year_of_Birth','Year_of_Death','Lifespan','Instance','Gender','Country_of_citizenship','Occupation','Letters_Received','Times_Mentioned','Brief_Description']
# person_org_col = ['XML_ID','Wikidata_ID','Viaf_ID','NHRP_ID','Given_Name','Surname','Name','Year_of_Birth','Year_of_Death','Lifespan','Instance','Gender','Country_of_citizenship','Occupation','Letters_Received','Times_Mentioned','Brief_Description']
# work_file = pd.read_csv('Person_Register_Info.csv', names=colnames, na_filter=False)
#
# ids_persons = work_file.XML_ID.tolist()
# persons_names = work_file.Name.tolist()


def count_letters_to_person(reg_entries):
    n_letters_received = {'XML_ID': 'Letters_Received'}
    mentions_list = {'XML_ID': 'Times_Mentioned'}
    for entry in reg_entries:
        xml_id = entry.attrib['{http://www.w3.org/XML/1998/namespace}id']
        received = entry.xpath('.//tei:div/tei:list[@type="recipient_letter"]/tei:item', namespaces=ns)
        mentions = entry.xpath('.//tei:div/tei:list[@type="mentioned_letter"]/tei:item', namespaces=ns)

        n_letters_received[xml_id] = (len(received))
        mentions_list[xml_id] = (len(mentions))

    return n_letters_received, mentions_list





# print(n_letter_received)
# print(n_mentions)


def edit_person_org_csv():
    colnames = ['XML_ID', 'Wikidata_ID', 'Viaf_ID', 'NHRP_ID', 'Given_Name', 'Surname', 'Name', 'Year_of_Birth',
                'Year_of_Death', 'Lifespan', 'Country_of_citizenship', 'Occupation', 'Instance', 'Gender',
                'Brief_Description']
    prof_col = ['Norsk','English']
    work_file = pd.read_csv('final_pers_org_details.csv', names=colnames, na_filter=False)
    prof_file = pd.read_csv('occupations_list_edited.csv',names=prof_col, na_filter=False)

    xmlID = work_file.XML_ID.tolist()
    wikidataID = work_file.Wikidata_ID.tolist()
    viafID = work_file.Viaf_ID.tolist()
    nhrpID = work_file.NHRP_ID.tolist()
    givenName = work_file.Given_Name.tolist()
    surname = work_file.Surname.tolist()
    name = work_file.Name.tolist()
    yearBirth = work_file.Year_of_Birth.tolist()
    yearDeath = work_file.Year_of_Death.tolist()
    lifespan = work_file.Lifespan.tolist()
    citizenship = work_file.Country_of_citizenship.tolist()
    occupation = work_file.Occupation.tolist()
    instance = work_file.Instance.tolist()
    gender = work_file.Gender.tolist()
    description = work_file.Brief_Description.tolist()
    occup_norsk = prof_file.Norsk.tolist()
    occup_eng = prof_file.English.tolist()

    instance_label = []
    for label in instance:
        if label not in instance_label:
            if label == str(''):
                continue
            else:
                instance_label.append(label)
    #pprint.pprint(instance_label)

    prof_label = []
    for prof in occupation:
        if prof not in prof_label:
            if prof == str(''):
                continue
            else:
                prof_label.append(prof)
    #pprint.pprint(prof_label)

    dict_instance_label = {'bank': 'bank',
                           'stiftet': 'foundation',
                           'embetskontor': 'civil service office',
                           'forening': 'society',
                           'studentforening': 'student society',
                           'studentforeningsstyre': 'student society board',
                           'bokhandel': 'bookshop',
                           'forlag': 'publisher',
                           'bokhandel og forlag': 'bookshop, publisher',
                           'organisasjon': 'organization',
                           'universitetsstyre': 'university board',
                           'litterært selskap': 'literary society',
                           'teater': 'theatre',
                           'interesseorganisasjon': 'interest group',
                           #'interesseorganisasjon og teateragentur': 'interest group, theatre agency',
                           'bedriftsstyre': 'company board',
                           'byrett': 'court',
                           'departement': 'ministry',
                           'fotograf': 'photographer',
                           'hotell': 'hotel',
                           'komit': 'commitee',
                           'statsoverhode': 'head of state',
                           #'teater, opera og ballett': 'theatre, opera, ballet',
                           'ballett':'ballet',
                           'arbeidersamfunn': 'writers’ organization',
                           'etat': 'budgetary',
                           'tidsskrift': 'magazine',
                           'forretningsfører': 'society head',
                           'hoffembete': '',
                           'bedrift': 'company',
                           'lokaladministrativ enhet': 'local administrative unit',
                           'fetevareforretning': ''}

    titles = {'fyrst':'Prince',
              'prins':'Prince',
              'dronning':'Queen',
              'keiser': 'Emperor',
              'keiserinne': 'Empress',
              'prinsesse': 'Princess',
              'greve': 'Count',
              'sultan': 'Sultan',
              'emir': 'Commander/Prince',
              'grevinne':'Countess',
              'kong':'King'}

    gender_indicators = [
        'hustru',
        'mor',
        'datter',
        'prinsesse',
        'dronning',
        'inne']

    nationality = {
        'norsk':'Norway',
        'dansk':'Denmark',
        'svensk':'Sweden',
        'tysk':'Germany',
        'østerriksk': 'Austria',
        'sveitsisk':'Switzerland',
        'engelsk':'England',
        'italiensk':'Italy',
        'spansk':'Spain',
        'fransk':'France',
        'gresk':'Greece',
        'polsk':'Poland',
        'amerikansk':'United States of America',
        'russisk':'Russia',
        'tsjekkisk':'Czechoslovakia',
        'britisk': 'United Kingdom of Great Britain and Ireland',
        'osmansk': 'Ottoman Empire'
    }

    female_profs = {
        'sangerinne': 'singer',
        'salongvertinne':'salonnière',
        'kvinnesaksforkjemper':'suffragette',
        'friherrinne':'Baroness',
        'skuespillerinne':'actor'
    }

    dict_occup = {}
    for line in range(1, 229):
        norsk = occup_norsk[line]
        eng = occup_eng[line]
        dict_occup[norsk] = eng


    new_inst_list = ['Instance']
    new_occup_list = ['Occupation']
    new_gender_list = ['Gender']
    new_nationality_list = ['Country_of_citizenship']
    for i in range(1, 1704):
        inst = instance[i]
        desc = description[i]
        professions = occupation[i]
        gen = gender[i]
        citi_ship = citizenship[i]
        xmlid = xmlID[i]

        # add missing instances
        if inst == str(''):
            key_list = []

            if 'org' in xmlid:
                key_list.append('organization')
                #print(dict_instance_label[key])
            #print(key_list)
            # if len(key_list) == 0:
            #     new_inst_list.append(str(''))
            # # elif len(key_list) == 1:
            # #     new_inst_list.append(key_list[0])
            # else:
            new_inst_list.append(key_list)
        else:
            inst_list = []
            if inst == 'human':
                inst_list.append(inst)
                for title in titles:
                    if title in desc:
                        inst_list.append(titles[title])
                    else:
                        continue

            elif 'org' in xmlid:
                inst_list.append('organization')
                inst_list.append(inst)

            #pprint.pprint(inst_list)
            if 'Princess' in inst_list:
                inst_list.remove('Prince')

            if 'Empress' in inst_list:
                inst_list.remove('Emperor')

            new_inst_list.append(inst_list)

        # add missing occupations
        temp_occup_list = []
        # if professions != str(''):
        #     temp_occup_list.append(professions)

        if 'human' not in new_inst_list[i]:
            temp_occup_list.append('organization')
        else:
            for k in dict_occup:
                if k in desc:
                    if dict_occup[k] in temp_occup_list:
                        continue
                    else:
                        temp_occup_list.append(dict_occup[k])

            if 'literary historian' in temp_occup_list:
                for items in temp_occup_list:
                    if items == 'literarian':
                        temp_occup_list.remove(items)
                    if items == 'historian':
                        temp_occup_list.remove(items)
                    else:
                        continue

            if 'Princess' in temp_occup_list:
                for items in temp_occup_list:
                    if items == 'Prince':
                        temp_occup_list.remove(items)
                    else:
                        continue

            if 'Empress' in temp_occup_list:
                for items in temp_occup_list:
                    if items == 'Emperor':
                        temp_occup_list.remove(items)
                    else:
                        continue

            if 'married to' in temp_occup_list:
                if gen == 'male':
                    for item in temp_occup_list:
                        if item == 'married to':
                            temp_occup_list.remove(item)
                        else:
                            continue
                else:
                    for item in temp_occup_list:
                        if item != 'married to':
                            temp_occup_list.remove(item)
                        else:
                            continue
                    if len(temp_occup_list) >= 1:
                        for item in temp_occup_list:
                            if item != 'married to':
                                temp_occup_list.remove(item)
                    for key in female_profs:
                        if key in desc:
                            temp_occup_list.append(female_profs[key])


        new_occup_list.append(temp_occup_list)

        # add missing genders
        new_occup = new_occup_list[i]
        if gen != str(''):
            new_gender_list.append(gen)
        else:
            if 'organization'in new_occup:
                new_gender_list.append(str(''))
            else:
                if any(indicator in desc for indicator in gender_indicators):
                    new_gender_list.append('female')
                else:
                    new_gender_list.append('male')

        # add missing citizenship information
        # !Overwrite the entries from Wikidata to normalise the different
        # names for a country, e.g. Nazi Germany -> Germany
        # if citi_ship == str(''):
        nat_list = []
        if 'hustru' in desc:
            nat_list.append(str(''))
        else:
            for k in nationality:
                if k in desc:
                    nat_list.append(nationality[k])
        new_nationality_list.append(nat_list)
        # else:
        #     new_nationality_list.append(citi_ship)
        # print(nat_list)





    for x in range(len(new_nationality_list)):
        new_input = re.sub('\[', '', re.sub('\]', '', re.sub('\'', '',str(new_nationality_list[x]))))
        new_nationality_list[x]=new_input

    for y in range(len(new_inst_list)):
        new_input = re.sub('\[', '', re.sub('\]', '', re.sub('\'', '',str(new_inst_list[y]))))
        new_inst_list[y]=new_input

    for z in range(len(new_occup_list)):
        new_input = re.sub('\[', '', re.sub('\]', '', re.sub('\'', '',str(new_occup_list[z]))))
        new_occup_list[z]=new_input
    # pprint.pprint(new_nationality_list)
    print(len(new_nationality_list))
    #pprint.pprint(new_nationality_list)
    #print(len(new_nationality_list))


    return new_inst_list,new_gender_list,new_nationality_list, new_occup_list


n_letter_received, n_mentions = count_letters_to_person(regest_entries)
#print(n_letter_received)
new_i_list, new_g_list,new_n_list,new_occup_list = edit_person_org_csv()


def extract_marriages(person_entries):
    colnames = ['XML_ID', 'Wikidata_ID', 'Viaf_ID', 'NHRP_ID', 'Given_Name', 'Surname', 'Name', 'Year_of_Birth',
                'Year_of_Death', 'Lifespan', 'Country_of_citizenship', 'Occupation', 'Instance', 'Gender',
                'Brief_Description']
    wf = pd.read_csv('final_pers_org_details.csv', names=colnames, na_filter=False)

    ids_persons = wf.XML_ID.tolist()
    persons_names = wf.Name.tolist()

    pers_entries = 0
    married_to_dict = {}
    person_id = []
    spouse_id = []
    for person in person_entries:
        xml_id = person.attrib['{http://www.w3.org/XML/1998/namespace}id']

        pers_entries += 1

        long_desc = person.xpath('./tei:p[text()[contains(.,"gift med")]]', namespaces=ns)
        for desc in long_desc:
            married_to_dict[xml_id] = {}

            def replace_all(text, dic):
                for i, j in dic.items():
                    text = text.replace(i, j)
                return text

            replace_string = {
                '\'': '',
                r'\n': '',
                '&#229;': 'å',
                '&#246;': 'ö',
                '&#248;': 'ø',
                '&#252;': 'ü',
                '&amp;': '&',
                '&#197;': 'Å'}

            person_text = str(etree.tostring(desc))
            person_texts = ' '.join(person_text.split())
            cleaned_pers_texts = replace_all(person_texts, replace_string)
            # print(cleaned_pers_texts)

            result = re.search(r'(?<=gift med )[^.]*', cleaned_pers_texts)

            try:
                ref_pers = re.search('(?<=HIS:hisRef type="person" target=")[^">]*', str(result))
                if ref_pers != None:
                    # print(ref_pers.group(0))
                    married_to_dict[xml_id]['married to'] = ref_pers.group(0)
                    spouse_id.append(ref_pers.group(0))
                    person_id.append(xml_id)
                else:
                    spouse_id.append(str(''))
                    person_id.append(xml_id)

            except TypeError:
                spouse_id.append(str(''))
                person_id.append(xml_id)
                continue

    # rows = zip(person_id, spouse_id)
    # with open('married_persons.csv', 'w', ) as work_csv:
    #     wr = csv.writer(work_csv, delimiter=',')
    #     for row in rows:
    #         wr.writerow(row)

    workbook = openpyxl.load_workbook('./married_persons.xlsm')
    sheet = workbook.active
    Person_id = []
    Spouse_id = []
    Spouse_Name = []

    for rowNum in range(2, 243):
        Person_id.append(sheet.cell(row=rowNum, column=1).value)
        Spouse_id.append(sheet.cell(row=rowNum, column=3).value)
        Spouse_Name.append(sheet.cell(row=rowNum, column=5).value)

    # print(Person_id)
    # print(Spouse_id)
    # print(Spouse_Name)
    # ids = person.attrib['{http://www.w3.org/XML/1998/namespace}id']

    for x, entry in enumerate(Spouse_id):
        if entry == None:
            Spouse_id[x] = str('')

    for x, entry in enumerate(Spouse_Name):
        if entry == None:
            Spouse_Name[x] = str('')

    new_married_to_list = ['Spouse\'s_Name']
    new_married_to_id_list = ['Spouse\'s_ID']
    #print(ids_persons)
    print(len(ids_persons))
    for n in range(1, 1704):
        id = ids_persons[n]
        # for id in ids_persons:

        if id in Person_id:
            i = Person_id.index(id)
            if Spouse_id[i] != '' and Spouse_Name[i] != '':
                new_married_to_id_list.append(Spouse_id[i])
                new_married_to_list.append(Spouse_Name[i].upper())
                # print(Spouse_id[i])
                # print(Spouse_Name[i])
            elif Spouse_id[i] != '' and Spouse_Name[i] == '':
                new_married_to_id_list.append(Spouse_id[i])
                id_tokens = Spouse_id[i].split(', ')
                # intermediate_id_list = []
                intermediate_name_list = []
                # print(Spouse_id[i])
                # print(id_tokens)
                for token in id_tokens:
                    if token in ids_persons:
                        n = ids_persons.index(token)
                        intermediate_name_list.append(persons_names[n])
                        # print(persons_names[n])
                    else:
                        intermediate_name_list.append(str(''))
                new_married_to_list.append(str(intermediate_name_list)[1:-1])
            elif Spouse_id[i] == '' and Spouse_Name[i] != '':
                new_married_to_list.append(Spouse_Name[i].upper())
                if Spouse_Name[i] == 'Edvard, prins av Wales':
                    cleaned_name = re.sub(',', '', re.sub('\)', '', re.sub('\(', '', Spouse_Name[i])))
                else:
                    cleaned_name = re.sub('\)','',re.sub('\(','', Spouse_Name[i]))
                capital_names = cleaned_name.upper()
                individual_names = capital_names.split(', ')
                #print(individual_names)
                intermediate_id_list = []
                # intermediate_name_list = []
                for name in individual_names:
                    name_tokens = name.split(' ')
                    #print(name)
                    matched_ids = []
                    for value in persons_names:
                        v = re.sub('\]','',re.sub('\[','', value))
                        # print(v)
                        # for token in tokenize_name:
                        # if all token in dict_person_reg.values():
                        # print(dict_person_reg.values())
                        if all(x in v for x in name_tokens):
                            # print(list(dict_person_reg.keys())[list(dict_person_reg.values()).index(value)])
                            print(v)
                            index_matched_name = persons_names.index(value)
                            print(ids_persons[index_matched_name])
                            matched_ids.append(ids_persons[index_matched_name])
                            #intermediate_id_list.append(ids_persons[index_matched_name])
                        else:
                            continue

                    if len(matched_ids) ==0:
                            intermediate_id_list.append(str('noID'))
                    else:
                        for match in matched_ids:
                            intermediate_id_list.append(match)
                new_married_to_id_list.append(intermediate_id_list)
            else:
                new_married_to_list.append(str(''))
                new_married_to_id_list.append(str(''))
        else:
            new_married_to_id_list.append(str(''))
            new_married_to_list.append(str(''))

    for x, item in enumerate(new_married_to_id_list):
        new_item = re.sub('\'','', re.sub('\[', '', re.sub('\]', '',str(item))))
        new_married_to_id_list[x] = new_item

    for x, item in enumerate(new_married_to_list):
        new_item = re.sub('\'', '', str(item))
        new_married_to_list[x] = new_item

    # pprint.pprint(married_to_dict)
    # print(Spouse_id)
    print(new_married_to_id_list)
    print(len(new_married_to_id_list))
    print(new_married_to_list)
    print(len(new_married_to_list))
    # print(len(ids_persons))

    return new_married_to_list, new_married_to_id_list


new_married_to_list, new_married_to_id_list = extract_marriages(persons_regest_entries)


def combile_to_csv(instance, gender,citizenship,occupation,n_letter_received, n_mentions, new_married_to_list, new_married_to_id_list):
    colnames = ['XML_ID', 'Wikidata_ID', 'Viaf_ID', 'NHRP_ID', 'Given_Name', 'Surname', 'Name', 'Year_of_Birth',
                'Year_of_Death', 'Lifespan', 'Country_of_citizenship', 'Occupation', 'Instance', 'Gender',
                'Brief_Description']
    work_file = pd.read_csv('final_pers_org_details.csv', names=colnames, na_filter=False)

    xmlID = work_file.XML_ID.tolist()
    wikidataID = work_file.Wikidata_ID.tolist()
    viafID = work_file.Viaf_ID.tolist()
    nhrpID = work_file.NHRP_ID.tolist()
    givenName = work_file.Given_Name.tolist()
    surname = work_file.Surname.tolist()
    name = work_file.Name.tolist()
    yearBirth = work_file.Year_of_Birth.tolist()
    yearDeath = work_file.Year_of_Death.tolist()
    lifespan = work_file.Lifespan.tolist()
    description = work_file.Brief_Description.tolist()
    n_l_received = []
    n_mention = []

    for key in n_letter_received:
        n_l_received.append(n_letter_received[key])

    for keys in n_mentions:
        n_mention.append(n_mentions[keys])

    rows = zip(xmlID,wikidataID,viafID,nhrpID,givenName,surname,name,yearBirth,yearDeath,lifespan,instance, gender,citizenship,occupation,n_l_received, n_mention,description, new_married_to_list, new_married_to_id_list)
    with open('Person_Register_Info.csv', 'w', ) as work_csv:
        wr = csv.writer(work_csv, delimiter=',')

        for row in rows:
            wr.writerow(row)


combile_to_csv(new_i_list, new_g_list,new_n_list,new_occup_list, n_letter_received, n_mentions, new_married_to_list, new_married_to_id_list)