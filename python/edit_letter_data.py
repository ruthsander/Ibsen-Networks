import xml.etree.ElementTree as ET
from lxml import etree
from html.entities import codepoint2name
import pprint
import re
import csv
from pywikibot.data import api
import pywikibot
import requests
from pywikibot import pagegenerators as pg
import pandas as pd
from csv import writer
from csv import reader
import time
import numpy as np
import sys
import openpyxl
import time
from geopy.geocoders import GeoNames

### This script edits the data collected using 'ExtractLetterData.py" ###


def edit_letter_info_csv():
    colnames = ['Number', 'Letter_ID','Sender_ID','Sender_Name','Recipient_ID','Recipient_Name','Date','Dispatch_Location','Dispatch_Location_Abbr','Mentioned_Person_ID','Mentioned_Person','Mentioned_Org_ID','Mentioned_Org','Mentioned_Places_Abbr','Mentioned_Places','Recipient_Adress', 'Mentioned_Works_IDs']
    person_org_col = ['XML_ID', 'Wikidata_ID', 'Viaf_ID', 'NHRP_ID', 'Given_Name', 'Surname', 'Name', 'Year_of_Birth',
                'Year_of_Death', 'Lifespan', 'Country_of_citizenship', 'Occupation', 'Instance', 'Gender',
                'Brief_Description']
    work_file = pd.read_csv('merged_letter_info.csv', names=colnames, na_filter=False)
    pers_org_file = pd.read_csv('final_pers_org_details.csv',names=person_org_col, na_filter=False)

    letter_id = work_file.Letter_ID.tolist()
    s_id = work_file.Sender_ID.tolist()
    s_name = work_file.Sender_Name.tolist()
    recip_id = work_file.Recipient_ID.tolist()
    recip_name = work_file.Recipient_Name.tolist()
    date = work_file.Date.tolist()
    disp_location = work_file.Dispatch_Location.tolist()
    disp_loc_abbr = work_file.Dispatch_Location_Abbr.tolist()
    ment_per_id = work_file.Mentioned_Person_ID.tolist()
    ment_pers = work_file.Mentioned_Person.tolist()
    ment_org_abbr = work_file.Mentioned_Org_ID.tolist()
    ment_org = work_file.Mentioned_Org.tolist()
    ment_pla_abbr = work_file.Mentioned_Places_Abbr.tolist()
    ment_place = work_file.Mentioned_Places.tolist()
    recip_addr = work_file.Recipient_Adress.tolist()
    ment_work_id = work_file.Mentioned_Works_IDs.tolist()

    # data list from person register file
    register_pid = pers_org_file.XML_ID.tolist()
    register_name = pers_org_file.Name.tolist()
    # print(len(register_name))
    dict_person_reg = {}
    for x in range(1,1704):
        dict_person_reg[register_pid[x]]= register_name[x]

    # create lists with information found in Excel file 'Works_ids_and_text_NME'
    #pprint.pprint(dict_person_reg)
    #print(len(ment_pers))
    workbook = openpyxl.load_workbook('../xml-filer/Works_ids_and_text_NME.xlsx')
    sheet = workbook.active
    verkID = []
    tittel = []
    tekstID = []
    sjanger = []
    for rowNum in range(2, 70):
        verkID.append(sheet.cell(row=rowNum, column=1).value)
        tittel.append(sheet.cell(row=rowNum, column=3).value)
        tekstID.append(sheet.cell(row=rowNum, column=4).value)
        sjanger.append(sheet.cell(row=rowNum, column=5).value)

    # create new lists for edited information
    new_sender_names = ['Sender_Name']
    new_mentioned_pers_names = ['Mentioned_Persons']
    new_mentioned_orgs = ['Mentioned_Org']
    new_recipient_ids = ['Recipient_ID']
    new_works_titles = ['Mention_Works_Title']
    new_tekst_ids = ['Text_ID']
    new_works_genre = ['Mention_Works_Genre']
    collectiv_work = ['Collective_Vesions_ID']
    collective = {}

    # create a dictionary with recipient's name as the key and percipient's id as the value
    recipient_dict = {}
    for n, m in zip(recip_name, recip_id):
        if n != '' and m != '':
            recipient_dict[n] = str(m)
        else:
            continue

    # for each letter in collection
    for i in range(1,2410):
        # clean sender name entries
        sender_name = ' '.join(s_name[i].split())
        # sender_id = s_id[i]
        if sender_name == 'IBSEN, HENRIK':
            sender_name = 'HENRIK IBSEN'
        new_sender_names.append(sender_name)

        # add mentioned persons' names
        m_per_id = re.sub('\'', '', str(ment_per_id[i]))
        #pprint.pprint(m_per_id)
        #print(m_per_id)
        intermediate_per_list = []
        split_ids = m_per_id.split(', ')
        if m_per_id == str(''):
            intermediate_per_list.append(str(''))
        else:
            for ids in split_ids:
                if ids in dict_person_reg:
                    intermediate_per_list.append(dict_person_reg[ids])
                else:
                    intermediate_per_list.append(str('-'))
                    #print(ids)
        new_mentioned_pers_names.append(str(intermediate_per_list).lstrip('[').rstrip(']'))

        # add mentioned org names
        intermediate_org_list = []
        ment_org_id = re.sub('\'', '', str(ment_org_abbr[i]))
        split_orgs = ment_org_id.split(', ')
        if ment_org_id == str(''):
            intermediate_org_list.append(str(''))
        else:
            for org in split_orgs:
                if org in dict_person_reg:
                    intermediate_org_list.append(dict_person_reg[org])
                else:
                    intermediate_org_list.append(str('-'))
                    # print(org)
        new_mentioned_orgs.append(str(re.sub('\'', '', str(intermediate_org_list))).lstrip('[').rstrip(']'))

        # add recipient id based on name
        intermediate_id_list = []
        recipient_name = recip_name[i]
        recipient_uppercase = recipient_name.upper()
        recipient_id = recip_id[i]
        # print(recipient_name)
        if recipient_name == '':
            intermediate_id_list.append(str(''))
        else:
            if recipient_uppercase in recipient_dict.keys():
                intermediate_id_list.append(recipient_dict[recipient_uppercase])
            elif recipient_name == 'ukjent mottager' or recipient_name == 'UKJENT MOTTAGER':
                # new_recipient_ids.append('unknown_per')
                intermediate_id_list.append('unknown_person')
            elif recipient_name == 'ukjente mottagere':
                intermediate_id_list.append('unknown_persons')
            elif recipient_uppercase == 'JOHAN CHRISTIAN DAHL':
                intermediate_id_list.append('peJCDa')
            elif recipient_uppercase == 'FREDERIK HEGEL':
                intermediate_id_list.append('peFH')
            elif recipient_uppercase == 'PETER HANSEN':
                intermediate_id_list.append('pePHa')
            elif recipient_uppercase == 'FREDERIK BLICH':
                intermediate_id_list.append('peDrB')
            elif recipient_uppercase == 'DAHL':
                intermediate_id_list.append('peDahl')
            elif recipient_uppercase == 'CHRISTIAN HAMMER':
                intermediate_id_list.append('peCH')
            elif recipient_uppercase == 'EMIL JONAS':
                intermediate_id_list.append('peEJ')
            elif recipient_uppercase == 'AUGUST LINDBERG':
                intermediate_id_list.append('peAL')
            elif recipient_uppercase == 'BJØRN BJØRNSON':
                intermediate_id_list.append('peBBj')
            elif recipient_uppercase == 'LUDVIG AUBERT':
                intermediate_id_list.append('peLMBA')
            elif recipient_uppercase == 'JOHAN HJORT':
                intermediate_id_list.append('peJHj')
            elif recipient_uppercase == 'FREDRIK GRØN':
                intermediate_id_list.append('peAFG')
            elif recipient_uppercase == 'NILS VOGT':
                intermediate_id_list.append('peNVo')
            elif recipient_uppercase == 'CHRISTIAN SONTUM':
                intermediate_id_list.append('peChrSo')
            elif recipient_uppercase == 'ERIK LIE':
                intermediate_id_list.append('peELie')
            elif recipient_uppercase == 'LAURA FITINGHOFF':
                intermediate_id_list.append('peLFi')
            elif recipient_uppercase == 'LUDVIG DAAE':
                intermediate_id_list.append('peLuDa')
            elif recipient_uppercase == 'OSCAR WERGELAND':
                intermediate_id_list.append('peOWe')
            elif recipient_uppercase == 'KRISTIANIA ARBEIDERSAMFUND':
                intermediate_id_list.append('orgKrArb')
            elif recipient_uppercase == 'HENRIK JÆGER':
                intermediate_id_list.append('peHeJa')
            elif recipient_uppercase == 'CHRISTIAN HOSTRUP':
                intermediate_id_list.append('peJCH')
            elif recipient_uppercase == 'JOHANNES STEEN':
                intermediate_id_list.append('peJSte')
            elif recipient_uppercase == 'KNUD KNUDSEN':
                intermediate_id_list.append('peKKspr')
            elif recipient_uppercase == 'THOMAS KRAG':
                intermediate_id_list.append('peTPK')
            else:
                tokenize_name = recipient_uppercase.split(' ')
                for value in dict_person_reg.values():
                    # for token in tokenize_name:
                    # if all token in dict_person_reg.values():
                    # print(dict_person_reg.values())
                    if all(x in value for x in tokenize_name):
                        # print(list(dict_person_reg.keys())[list(dict_person_reg.values()).index(value)])
                        matched_id = list(dict_person_reg.keys())[list(dict_person_reg.values()).index(value)]
                        intermediate_id_list.append(matched_id)
        new_recipient_ids.append(intermediate_id_list)
    #     print(intermediate_id_list)
    #     print(len(intermediate_id_list))
    #     print(tokenize_name)
    #     print(dict_person_reg.values())
    #
    #     for value in dict_person_reg.values():
    #         print(value)
    #     for token in tokenize_name:
    #         for value in dict_person_reg.values():
    #             if not token in value:
    #                 continue
    #             else:
    #                 new_recipient_ids.append(dict_person_reg)
    #
    # pprint.pprint(new_recipient_ids)
    #         for a in list:
    #             if not a in str:
    # pprint.pprint(new_mentioned_orgs)
    # print(new_mentioned_orgs[123])
    #print(new_sender_names)
    # pprint.pprint(new_recipient_ids)
    # print(len(new_recipient_ids))


        # return new_recipient_ids, recip_name

        # match works info to ids
        men_w_id = ment_work_id[i]
        intermediate_works_title = []
        intermediate_tekst_id = []
        intermediate_works_genre = []
        #print(men_w_id)
        if men_w_id == '':
            intermediate_works_title.append(str(''))
            intermediate_tekst_id.append(str(''))
            intermediate_works_genre.append(str(''))
        else:
            clean_work_id = re.sub('\'', '', str(men_w_id))
            work_tokens = clean_work_id.split(', ')
            for t_w in work_tokens:
                if any(c in t_w for c in ('(',')')):
                    t_w_cleaned = re.sub('\(', '',(re.sub('\)', '',str(t_w))))
                    #print(t_w_cleaned)
                    if t_w_cleaned in verkID:
                        index = verkID.index(t_w_cleaned)
                        intermediate_works_title.append(str('(')+tittel[index]+str(')'))
                        intermediate_tekst_id.append(str('(')+tekstID[index]+str(')'))
                        intermediate_works_genre.append(str('(')+sjanger[index]+str(')'))
                    else:
                        intermediate_works_title.append(t_w_cleaned)
                        intermediate_tekst_id.append(t_w_cleaned)
                        intermediate_works_genre.append(t_w_cleaned)
                else:
                    if t_w in verkID:
                        index = verkID.index(t_w)
                        intermediate_works_title.append(tittel[index])
                        intermediate_tekst_id.append(tekstID[index])
                        intermediate_works_genre.append(sjanger[index])
                    else:
                        intermediate_works_title.append(t_w)
                        intermediate_tekst_id.append(t_w)
                        intermediate_works_genre.append(t_w)
        new_works_titles.append(intermediate_works_title)
        new_tekst_ids.append(intermediate_tekst_id)
        new_works_genre.append(intermediate_works_genre)

    #pprint.pprint(ment_work_id)
    #pprint.pprint(new_works_titles)
    # pprint.pprint(len(new_works_genre))
    # pprint.pprint(len(new_tekst_ids))
    # pprint.pprint(new_tekst_ids)
    # pprint.pprint(new_works_genre)
    # pprint.pprint(len(new_works_genre))

    #add collective Id for versions of a work
    for text_id in new_tekst_ids[1:2410]:
        #print(text_id)
        intermediate_collective = []
        if text_id == '':
            intermediate_collective.append(str(''))
        else:
            text_tokens = str(text_id).split(', ')
            try:
                for t_tokens in text_tokens:
                    if any(c in t_tokens for c in('C1', 'C2')):
                        intermediate_collective.append('C')
                    elif any(c in t_tokens for c in('F1', 'F2')):
                        intermediate_collective.append('F')
                    elif any(c in t_tokens for c in('H1', 'H2')):
                        intermediate_collective.append('HH')
                    elif any(c in t_tokens for c in('G1', 'G2')):
                        intermediate_collective.append('G')
                    elif any(c in t_tokens for c in('K1', 'K2')):
                        intermediate_collective.append('K')
                    elif any(c in t_tokens for c in('OL1', 'OL2','OL3')):
                        intermediate_collective.append('OL')
            except None:
                intermediate_collective.append(str(''))

        remove_duplicates_list = list(dict.fromkeys(intermediate_collective))
            # clean_titles = re.sub('\[', '', re.sub('\]', '', str(remove_duplicates)))
            # collectiv_work[x] = remove_duplicates
        collectiv_work.append(remove_duplicates_list)


    # pprint.pprint(collectiv_work)
    # print(len(collectiv_work))

    ### Clean lists ###
    for x, item in enumerate(new_works_titles):
        new_item = re.sub('\(', '[', re.sub('\)', ']', re.sub('\'', '', re.sub('\[', '', re.sub('\]', '', str(item))))))
        new_works_titles[x] = new_item

    for x, item in enumerate(ment_work_id):
        new_item = re.sub('\(', '[', re.sub('\)', ']', re.sub('\'', '', re.sub('\[', '', re.sub('\]', '',str(item))))))
        ment_work_id[x] = new_item

    for x, item in enumerate(new_works_genre):
        new_item = re.sub('\(', '[', re.sub('\)', ']',re.sub('\'', '', re.sub('\[', '', re.sub('\]', '',str(item))))))
        new_works_genre[x] = new_item

    for x, item in enumerate(collectiv_work):
        new_item = re.sub('\(', '[', re.sub('\)', ']',re.sub('\'','', re.sub('\[', '', re.sub('\]', '',str(item))))))
        collectiv_work[x] = new_item

    for x, item in enumerate(new_tekst_ids):
        new_item = re.sub('\(', '[', re.sub('\)', ']', re.sub('\'','', re.sub('\[', '', re.sub('\]', '',str(item))))))
        new_tekst_ids[x] = new_item

    for x, item in enumerate(ment_per_id):
        new_item = re.sub('\'','', re.sub('\[', '', re.sub('\]', '',str(item))))
        ment_per_id[x] = new_item

    for x, item in enumerate(new_mentioned_pers_names):
        new_item = re.sub('\'','', re.sub('\[', '', re.sub('\]', '',str(item))))
        new_mentioned_pers_names[x] = new_item

    for x, item in enumerate(new_recipient_ids):
        new_item = re.sub('\'','',re.sub('"','',  re.sub('\[', '', re.sub('\]', '',str(item)))))
        new_recipient_ids[x] = new_item

    #pprint(collectiv_work)
    # print(len(letter_id))
    # print(len(s_id))
    # print(len(new_sender_names))
    # print(len(new_recipient_ids))
    # print(len(recip_name))
    # print(len(date))
    # print(len(ment_place))
    # print(len(ment_pla_abbr))
    # print(len(ment_per_id))
    # print(len(new_mentioned_pers_names))
    # print(len(ment_org_abbr))
    # print(len(new_mentioned_orgs))
    # print(len(ment_work_id))
    # print(len(new_works_titles))
    # print(len(new_tekst_ids))
    # print(len(new_works_genre))
    # print(len(collectiv_work))
    #
    # print(new_works_titles)
    # print(ment_work_id)
    # print(new_tekst_ids)
    # print(new_works_genre)
    # print(collectiv_work)

    return letter_id,s_id,new_sender_names, new_recipient_ids, recip_name, date,ment_place, ment_pla_abbr, ment_per_id, new_mentioned_pers_names, ment_org_abbr, new_mentioned_orgs, ment_work_id, new_works_titles, new_tekst_ids, new_works_genre, collectiv_work


letter_id,s_id,new_sender_names, new_recipient_ids, recip_name, date,ment_place, ment_pla_abbr, ment_per_id, new_mentioned_pers_names, ment_org_abbr, new_mentioned_orgs, ment_work_id, new_works_titles, new_tekst_ids, new_works_genre, collectiv_work = edit_letter_info_csv()
# strangeId, strangeName = edit_letter_info_csv()
#edit_letter_info_csv()

# def recipIDsANDnames_csv(id,appel):
#
#     rows = zip(id,appel)
#     with open('recipIDsANDnames2.csv', 'w', ) as work_csv:
#         wr = csv.writer(work_csv, delimiter=',')
#
#         for row in rows:
#             wr.writerow(row)
#
#
# #recipIDsANDnames_csv(strangeId,strangeName)


def geo_data():
    colnames = ['Letter_ID','Sender_ID','Sender_Name','Recipient_ID','Recipient_Name','Date','Dispatch_Location','Dispatch_Location_Abbr','Mentioned_Person_ID','Mentioned_Person','Mentioned_Org_ID','Mentioned_Org','Mentioned_Places_Abbr','Mentioned_Places','Recipient_Adress', 'Mentioned_Works_IDs']

    work_file = pd.read_csv('merged_letter_info.csv', names=colnames, na_filter=False)

    disp_loc_abbr = work_file.Dispatch_Location_Abbr.tolist()
    disp_location = work_file.Dispatch_Location.tolist()
    ment_loc_abbr = work_file.Mentioned_Places_Abbr.tolist()
    ment_location = work_file.Mentioned_Places.tolist()
    recip_address = work_file.Recipient_Adress.tolist()
    # disp_location = list(dict.fromkeys(disp_location))
    # print(disp_location)

    # create list of all unique locations in mentioned places, add these to dictionary
    ment_place_list = []
    ment_ab_list = []
    places_dict = {}
    for i in range(1, 2410):
        y = ment_location[i]
        z = ment_loc_abbr[i]
        if y == '':
            continue
        else:
            clean_y = re.sub('\'', '',re.sub('"', '',str(y)))
            clean_z = re.sub('\'', '',re.sub('"', '',str(z)))
            tokens = clean_y.split(', ')
            t_abbr = clean_z.split(', ')
            for n, m in zip(tokens,t_abbr):
                if n not in places_dict.keys():
                    places_dict[n] = m
                else:
                    continue
                if n not in ment_place_list:
                    ment_ab_list.append(m)
                    ment_place_list.append(n)
                else:
                    continue

    # pprint.pprint(places_dict)

    # capitalize place names
    for x, item in enumerate(disp_location):
        item = ' '.join(item.split())
        capitaized_item = item.title()
        disp_location[x] = capitaized_item

    #full_abbr_dict = dict(zip(disp_loc_abbr, disp_location))

    # create dictionary of all unique dispatch location
    locations_dict = {}
    loc_list = []
    abbr_list = []
    for i in range(1, 2410):
        if disp_location[i] == '':
            continue
        elif disp_location[i] not in locations_dict.keys():
            locations_dict[disp_location[i]] = {}
            locations_dict[disp_location[i]]['Abbreviation'] = disp_loc_abbr[i]
           # locations_dict[disp_loc_abbr[i]].append(disp_location[i])
        # replace dictionary value if previously empty of equal to 'NN'
        elif locations_dict[disp_location[i]] == '' or locations_dict[disp_location[i]] == 'NN':
            # locations_dict[disp_location[i]] = dict.disp_loc_abbr[i]
            locations_dict[disp_location[i]]['Abbreviation'] = disp_loc_abbr[i]
        else:
            continue
            # print(locations_dict[disp_loc_abbr[i]])
            # for vlaues in locations_dict[disp_loc_abbr]:
            #     if disp_location in vlaues:
            #         continue
            #     else:
            #         locations_dict[disp_loc_abbr].append(disp_location[i])

    ### ADD Locations mentioned to locations_dicts ###
    # print(len(ab_list))
    # print(len(place_list))
    # places_dict = {}
    # for d in range(0,394):
    #
    #     if place_list[d] not in locations_dict.keys():
    #         locations_dict[place_list[d]]={}
    #         locations_dict[place_list[d]]['Abbreviation'] = ab_list[d]
    #     else:
    #         continue
    #         # except KeyError:
    #         #     print(k)


    ### GET GEONAMES INFO ###
    for key in locations_dict.keys():
        if key == 'Pillnitz Ved Dresden' or key == '[Pillnitz Ved Dresden]':
            key_edit = 'Pillnitz'
            try:
                gn = GeoNames(username='mastersstudent')  # enter GeoNames username
                place = (gn.geocode(key_edit))
                raw_data = place.raw
                locations_dict[key]['toponymName'] = raw_data['toponymName']
                locations_dict[key]['geonameId'] = raw_data['geonameId']
                locations_dict[key]['lat'] = raw_data['lat']
                locations_dict[key]['lng'] = raw_data['lng']
                locations_dict[key]['countryName'] = raw_data['countryName']
                locations_dict[key]['countryId'] = raw_data['countryId']
            except AttributeError:
                print(key)
        elif key == 'Amalfi':
            key_edit = 'Amalfi, Italy'
            try:
                gn = GeoNames(username='mastersstudent')
                place = (gn.geocode(key_edit))
                raw_data = place.raw
                locations_dict[key]['toponymName'] = raw_data['toponymName']
                locations_dict[key]['geonameId'] = raw_data['geonameId']
                locations_dict[key]['lat'] = raw_data['lat']
                locations_dict[key]['lng'] = raw_data['lng']
                locations_dict[key]['countryName'] = raw_data['countryName']
                locations_dict[key]['countryId'] = raw_data['countryId']
            except AttributeError:
                print(key)
        elif key == 'San Germano':
            key_edit = 'San Germano, Italy'
            try:
                gn = GeoNames(username='mastersstudent')
                place = (gn.geocode(key_edit))
                raw_data = place.raw
                locations_dict[key]['toponymName'] = raw_data['toponymName']
                locations_dict[key]['geonameId'] = raw_data['geonameId']
                locations_dict[key]['lat'] = raw_data['lat']
                locations_dict[key]['lng'] = raw_data['lng']
                locations_dict[key]['countryName'] = raw_data['countryName']
                locations_dict[key]['countryId'] = raw_data['countryId']
            except AttributeError:
                print(key)

        elif key == '[Fredrikshavn':
            key_edit = 'Frederikshavn, Denmark'
            try:
                gn = GeoNames(username='mastersstudent')
                place = (gn.geocode(key_edit))
                raw_data = place.raw
                locations_dict[key]['toponymName'] = raw_data['toponymName']
                locations_dict[key]['geonameId'] = raw_data['geonameId']
                locations_dict[key]['lat'] = raw_data['lat']
                locations_dict[key]['lng'] = raw_data['lng']
                locations_dict[key]['countryName'] = raw_data['countryName']
                locations_dict[key]['countryId'] = raw_data['countryId']
            except AttributeError:
                print(key)

        elif key == 'Kitzbüchel, Tyrol':
            key_edit = 'Kitzbühel'
            try:
                gn = GeoNames(username='mastersstudent')
                place = (gn.geocode(key_edit))
                raw_data = place.raw
                locations_dict[key]['toponymName'] = raw_data['toponymName']
                locations_dict[key]['geonameId'] = raw_data['geonameId']
                locations_dict[key]['lat'] = raw_data['lat']
                locations_dict[key]['lng'] = raw_data['lng']
                locations_dict[key]['countryName'] = raw_data['countryName']
                locations_dict[key]['countryId'] = raw_data['countryId']
            except AttributeError:
                print(key)
        elif key == 'Ala':
            key_edit = 'Ala, Italy'
            try:
                gn = GeoNames(username='mastersstudent')
                place = (gn.geocode(key_edit))
                raw_data = place.raw
                locations_dict[key]['toponymName'] = raw_data['toponymName']
                locations_dict[key]['geonameId'] = raw_data['geonameId']
                locations_dict[key]['lat'] = raw_data['lat']
                locations_dict[key]['lng'] = raw_data['lng']
                locations_dict[key]['countryName'] = raw_data['countryName']
                locations_dict[key]['countryId'] = raw_data['countryId']
            except AttributeError:
                print(key)
        else:
            try:
                gn = GeoNames(username='mastersstudent')
                place = (gn.geocode(key))
                raw_data = place.raw
                locations_dict[key]['toponymName'] = raw_data['toponymName']
                locations_dict[key]['geonameId'] = raw_data['geonameId']
                locations_dict[key]['lat'] = raw_data['lat']
                locations_dict[key]['lng'] = raw_data['lng']
                locations_dict[key]['countryName'] = raw_data['countryName']
                locations_dict[key]['countryId'] = raw_data['countryId']

                time.sleep(10)  # may work with shorter buffer

            except AttributeError:
                print(key)
    #
    # #pprint.pprint(locations_dict)
    # # print(locations_dict.keys())

    # create new lists for complied data from GeoNames
    correct_abbr = ['Abbriviations']
    geonameID_list = ['GeoName_ID']
    toponymName_list = ['toponymName']
    countryName_list = ['Country']
    latitude_list = ['Latitude']
    longitude_list = ['Longitude']
    new_disp_location_list = []
    new_disp_abbr_locations = []
    new_recipient_addr_list = ['Recipient_Location']
    new_recipient_addr_abbr = ['Recipient_Location_Abbr']
    recipiet_latitude = ['Recipient_Addr_Latitude']
    recipiet_longitude = ['Recipient_Addr_Longitude']

    # collect missing information by looking up location names/ abbreviations in dictionary
    for full, abbr in zip(disp_location, disp_loc_abbr):
        if full == '' and abbr == '':
            new_disp_abbr_locations.append(str(''))
            new_disp_location_list.append(str(''))
        elif full == '' and abbr != '':
            new_disp_abbr_locations.append(abbr)
            for keys in locations_dict.keys():
                if locations_dict[keys]['Abbreviation'] == abbr:
                    new_disp_location_list.append(keys)
                else:
                    continue
        elif full != '' and abbr == '':
            new_disp_location_list.append(full)
            if full in locations_dict:
                new_disp_abbr_locations.append(locations_dict[full]['Abbreviation'])
            else:
                continue
        else:
            new_disp_location_list.append(full)
            if abbr == 'NN':
                if full in locations_dict:
                    new_disp_abbr_locations.append(locations_dict[full]['Abbreviation'])
                else:
                    new_disp_abbr_locations.append(abbr)
            else:
                new_disp_abbr_locations.append(abbr)

    # print(new_disp_abbr_locations)
    # print(len(new_disp_abbr_locations))
    # print(new_disp_location_list)
    # print(len(new_disp_location_list))

    ### Match Place names with recipient adress ###
    for items in range(1, 2410):
        full_addr = recip_address[items]
        if full_addr == '':
            new_recipient_addr_list.append(str(''))
            new_recipient_addr_abbr.append(str(''))
        else:
            cleaned_addr = re.sub('\(', '', (
                re.sub('\)', '', re.sub('\'', '', re.sub('\.', '', re.sub(',', '', ' '.join(full_addr.split())))))))
            addr_tokens = cleaned_addr.split(' ')
            # print(addr_tokens)
            intermediate_address = []
            intermediate_abbr = []

            # match tokens with mentioned locations
            for token in addr_tokens:
                # print(token)
                if token == '':
                    continue
                else:
                    if token in ment_place_list:
                        # print(token)
                        index = ment_place_list.index(token)
                        intermediate_address.append(token)
                        intermediate_abbr.append(ment_ab_list[index])
                    elif token == 'Porsgrund':
                        intermediate_abbr.append(str(''))
                        intermediate_address.append(str('Porsgrund'))
                    elif token == 'Laurdal':
                        intermediate_abbr.append(str(''))
                        intermediate_address.append(str('lårdal'))
                    elif token == 'Bern':
                        intermediate_abbr.append(str(''))
                        intermediate_address.append(str('Bern'))
                    else:

                        intermediate_abbr.append(str(''))
                        intermediate_address.append(str(''))
            new_recipient_addr_list.append(intermediate_address)
            new_recipient_addr_abbr.append(intermediate_abbr)

    # clean found places from previous step
    for x, item in enumerate(new_recipient_addr_list):
        if x == 0:
            new_recipient_addr_list[x] = str(item)
        else:
            remove_empty_str = [s for s in item if s]
            item_cleaned = re.sub('\[', '', re.sub('\]', '', re.sub('\'', '', str(remove_empty_str))))
            if item_cleaned == 'Schweden, Norwegen, Wien':
                new_recipient_addr_list[x] = str('Wien')
            elif item_cleaned == 'Ungarn, Christiania, Norwegen':
                new_recipient_addr_list[x] = str('Anarcs, Ungarn')
            elif item_cleaned == 'Bergens, Bergen, Norwegen':
                new_recipient_addr_list[x] = str('Bergen, Norwegen')
            elif item_cleaned == 'Lund, Kristiania, Norwegen':
                new_recipient_addr_list[x] = str('Kristiania, Norwegen')
            else:
                new_recipient_addr_list[x] = str(item_cleaned)

    for x, item in enumerate(new_recipient_addr_abbr):
        if x == 0:
            new_recipient_addr_abbr[x] = str(item)
        else:
            remove_empty_str = [s for s in item if s]
            item_cleaned = re.sub('\[', '', re.sub('\]', '', re.sub('\'', '', str(remove_empty_str))))
            if item_cleaned == 'SE, N, Wien':
                new_recipient_addr_abbr[x] = str('Wien')
            elif item_cleaned == 'Ung, Kri, N':
                new_recipient_addr_abbr[x] = str('Ung')
            elif item_cleaned == 'Berg, Berg, N':
                new_recipient_addr_abbr[x] = str('Berg, N')
            elif item_cleaned == 'Lund, Kri, N':
                new_recipient_addr_abbr[x] = str('Kri, N')
            else:
                new_recipient_addr_abbr[x] = str(item_cleaned)

    print(ment_place_list)
    # print(new_recipient_addr_abbr)
    # print(len(new_recipient_addr_abbr))
    # print(new_recipient_addr_list)
    # print(len(new_recipient_addr_list))

    ### Sort through information from Geonames ###
    for r in range(1,2410):
        a = disp_location[r]
        b = new_recipient_addr_list[r]
        if a == '':
            correct_abbr.append(str(''))
            geonameID_list.append(str(''))
            countryName_list.append(str(''))
            toponymName_list.append(str(''))
            latitude_list.append(str(''))
            longitude_list.append(str(''))
        else:
            try:
                if a in locations_dict.keys():
                    # print(locations_dict[a])
                    try:
                        correct_abbr.append(str(locations_dict[a]['Abbreviation']))
                    except KeyError:
                        correct_abbr.append(str(''))
                    try:
                        geonameID_list.append(str(locations_dict[a]['geonameId']))
                    except KeyError:
                        geonameID_list.append(str(''))
                    try:
                        toponymName_list.append(str(locations_dict[a]['toponymName']))
                    except KeyError:
                        toponymName_list.append(str(''))
                    try:
                        countryName_list.append(str(locations_dict[a]['countryName']))
                    except KeyError:
                        countryName_list.append(str(''))
                    try:
                        latitude_list.append(str(locations_dict[a]['lat']))
                    except KeyError:
                        latitude_list.append(str(''))
                    try:
                        longitude_list.append(str(locations_dict[a]['lng']))
                    except KeyError:
                        longitude_list.append(str(''))

                else:
                    correct_abbr.append(str(''))
                    geonameID_list.append(str(''))
                    countryName_list.append(str(''))
                    toponymName_list.append(str(''))
                    latitude_list.append(str(''))
                    longitude_list.append(str(''))
            except ValueError:
                print(a)
        if b == '':
            recipiet_latitude.append(str(''))
            recipiet_longitude.append(str(''))
        else:
            if b in locations_dict.keys():
                recipiet_latitude.append(locations_dict[b]['lat'])
                recipiet_longitude.append(locations_dict[b]['lng'])
            else:
                try:
                    gn = GeoNames(username='mastersstudent')
                    place = (gn.geocode(b))
                    raw_data = place.raw

                    recipiet_latitude.append(str(raw_data['lat']))
                    recipiet_longitude.append(str(raw_data['lng']))

                    time.sleep(10)

                except AttributeError:
                    recipiet_latitude.append(str(''))
                    recipiet_longitude.append(str(''))
                    # print(key)

    # print(len(new_disp_location_list))
    # print(len(new_disp_abbr_locations))
    # print(len(geonameID_list))
    # print(len(toponymName_list))
    # print(len(countryName_list))
    # print(len(latitude_list))
    # print(len(longitude_list))
    # print(len(ment_place_list))
    # print(len(ment_ab_list))
    # print(len(recip_address))
    # print(len(new_recipient_addr_list))
    # print(len(new_recipient_addr_abbr))
    # print(len(recipiet_latitude))
    # print(len(recipiet_longitude))
    # print(geonameID_list)
    # print(len(geonameID_list))
    # print(countryName_list)
    # print(len(countryName_list))
    # print(correct_abbr)
    # print(len(correct_abbr))
    # print(toponymName_list)
    # print(len(toponymName_list))
    # print(latitude_list)
    # print(len(latitude_list))
    # print(longitude_list)
    # print(len(longitude_list))
    # print(len(new_recipient_addr_list))
    # print(len(new_recipient_addr_abbr))
    # print(recipiet_longitude)
    # print(len(recipiet_longitude))
    # print(recipiet_latitude)
    # print(len(recipiet_latitude))

    ### Clean up lists###
    for x, item in enumerate(recip_address):
        new_item = re.sub('\'','', re.sub('\[', '', re.sub(',,', ',', re.sub('\]', '',str(item)))))
        cleaned = ' '.join(new_item.split())
        recip_address[x] = cleaned
    for x, item in enumerate(ment_location):
        new_item = re.sub('\'','', re.sub('\[', '', re.sub('\]', '',str(item))))
        ment_location[x] = new_item
    for x, item in enumerate(ment_loc_abbr):
        new_item = re.sub('\'','', re.sub('\[', '', re.sub('\]', '',str(item))))
        ment_loc_abbr[x] = new_item


    # pprint.pprint(recip_address)
    return new_disp_location_list, new_disp_abbr_locations, geonameID_list,toponymName_list, countryName_list, latitude_list, longitude_list,recip_address, new_recipient_addr_list, new_recipient_addr_abbr, recipiet_latitude, recipiet_longitude, ment_location, ment_loc_abbr


new_disp_location_list, new_disp_abbr_locations, geonameID_list,toponymName_list, countryName_list, latitude_list, longitude_list,recip_address, new_recipient_addr_list, new_recipient_addr_abbr, recipiet_latitude, recipiet_longitude, ment_location, ment_loc_abbr = geo_data()
# geo_data()


def compile_final_csv(letter_id, s_id, new_sender_names, new_recipient_ids, recip_name, date,new_disp_location_list, new_disp_abbr_locations, geonameID_list, toponymName_list, countryName_list,
        latitude_list, longitude_list,recip_address, new_recipient_addr_list,
        new_recipient_addr_abbr, recipiet_latitude, recipiet_longitude, ment_per_id, new_mentioned_pers_names,ment_org_abbr, new_mentioned_orgs,ment_work_id, new_works_titles, new_tekst_ids, new_works_genre,
        collectiv_work,  ment_location, ment_loc_abbr):


    rows = zip( letter_id, s_id, new_sender_names, new_recipient_ids, recip_name, date,new_disp_location_list, new_disp_abbr_locations, geonameID_list, toponymName_list, countryName_list,
        latitude_list, longitude_list,recip_address, new_recipient_addr_list,
        new_recipient_addr_abbr, recipiet_latitude, recipiet_longitude, ment_per_id, new_mentioned_pers_names,ment_org_abbr, new_mentioned_orgs,ment_work_id, new_works_titles, new_tekst_ids, new_works_genre,
        collectiv_work,  ment_location, ment_loc_abbr)

    with open('Compiled_Letter_Data.csv', 'w', ) as work_csv:
        wr = csv.writer(work_csv, delimiter=',')

        for row in rows:
            wr.writerow(row)


compile_final_csv(letter_id, s_id, new_sender_names, new_recipient_ids, recip_name, date,new_disp_location_list, new_disp_abbr_locations, geonameID_list, toponymName_list, countryName_list,
        latitude_list, longitude_list,recip_address, new_recipient_addr_list,
        new_recipient_addr_abbr, recipiet_latitude, recipiet_longitude, ment_per_id, new_mentioned_pers_names,ment_org_abbr, new_mentioned_orgs,ment_work_id, new_works_titles, new_tekst_ids, new_works_genre,
        collectiv_work,  ment_location, ment_loc_abbr)