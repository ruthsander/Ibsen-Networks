import xml.etree.ElementTree as ET
from lxml import etree
from html.entities import codepoint2name
import pprint
import re
import csv
from pywikibot.data import api
import pywikibot
import requests
from pywikibot import pagegenerators as pg
import pandas as pd
from csv import writer
from csv import reader
import time
import numpy as np
import sys
import openpyxl

### Collect all columns from file 'Compiler_Letter_Data.csv' ###
colnames = ['Letter_ID','Sender_ID','Sender_Name','Recipient_ID','Recipient_Name','Date',
            'Dispatch_Location','Dispatch_Location_Abbr','GeoName_ID','toponymName','Country',
            'Latitude','Longitude','Recipient_Adress','Recipient_Location','Recipient_Location_Abbr',
            'Recipient_Addr_Latitude','Recipient_Addr_Longitude','Mentioned_Person_ID',
            'Mentioned_Persons','Mentioned_Org_ID','Mentioned_Org','Mentioned_Works_IDs',
            'Mention_Works_Title','Text_ID','Mention_Works_Genre','Collective_Vesions_ID',
            'Mentioned_Places','Mentioned_Places_Abbr']
work_file = pd.read_csv('Compiled_Letter_Data.csv', names=colnames, na_filter=False)

Letter_ID = work_file.Letter_ID.tolist()
Mention_Works_Title = work_file.Mention_Works_Title.tolist()
Mentioned_Works_IDs = work_file.Mentioned_Works_IDs.tolist()
Mention_Works_Genre = work_file.Mention_Works_Genre.tolist()
Text_ID = work_file.Text_ID.tolist()
Collective_Vesions_ID = work_file.Collective_Vesions_ID.tolist()

### Read xml files ###
ns = {'HIS': 'http://www.example.org/ns/HIS',
      'xml': 'http://www.w3.org/XML/1998/namespace',
      'tei': 'http://www.tei-c.org/ns/1.0'}

tree1844_1871 = etree.parse('../github/Ibsen-Networks/xml-data/B1844-1871ht.xml')
tree1871_1879 = etree.parse('../github/Ibsen-Networks/xml-data/B1871-1879ht.xml')
tree1880_1889 = etree.parse('../github/Ibsen-Networks/xml-data/B1880-1889ht.xml')
tree1890_1905 = etree.parse('../github/Ibsen-Networks/xml-data/B1890-1905ht.xml')

### jump to text element in xml files###
letter_text1844_1871 = tree1844_1871.xpath('//tei:text[@rend="letter"]', namespaces=ns)
letter_text1871_1879 = tree1871_1879.xpath('//tei:text[@rend="letter"]', namespaces=ns)
letter_text1880_1889 = tree1880_1889.xpath('//tei:text[@rend="letter"]', namespaces=ns)
letter_text1890_1905 = tree1890_1905.xpath('//tei:text[@rend="letter"]', namespaces=ns)


# search through all letter text for the names of Ibsen's works and
# add the information to the appropriate lists
def ids_and_title_from_letters(letters1844_1871):
    all_corresp_ids = {}
    for corresp in letter_text1844_1871:
        corresp_id = corresp.attrib['corresp']
        all_corresp_ids[corresp_id]=[]
        letter_text = corresp.xpath('.//tei:div[@type="letter"]/tei:p/text()', namespaces=ns)
        # print(letter_text)
        mentioned_titles = []
        if 'Brand.' in str(letter_text) or 'Brand ' in str(letter_text) or 'Brand»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Brand')
        if 'Bygmester Solness' in str(letter_text):
            all_corresp_ids[corresp_id].append('Bygmester Solness')
        if 'Catilina' in str(letter_text):
            all_corresp_ids[corresp_id].append('Catilina')
        if 'De unges Forbund' in str(letter_text):
            all_corresp_ids[corresp_id].append('De unges Forbund')
        if 'En folkefiende' in str(letter_text):
            all_corresp_ids[corresp_id].append('En folkefiende')
        if 'Episk Brand' in str(letter_text):
            all_corresp_ids[corresp_id].append('Episk Brand')
        if 'Et dukkehjem' in str(letter_text):
            all_corresp_ids[corresp_id].append('Et dukkehjem')
        if 'Fjeldfuglen' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fjeldfuglen')
        if 'Fruen fra havet' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fruen fra havet')
        if 'Fru Inger til Østeraad' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fru Inger til Østeraad')
        if 'Fru Inger til Østråt' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fru Inger til Østråt')
        if 'Gengangere' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gengangere')
        if 'Gildet paa Solhoug' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gildet paa Solhoug')
        if 'Gildet på Solhaug' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gildet på Solhaug')
        if 'Hedda Gabler' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hedda Gabler')
        if 'Hvide heste' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hvide heste')
        if 'Hærmændene paa Helgeland' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hærmændene paa Helgeland')
        if 'Hærmændene på Helgeland' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hærmændene på Helgeland')
        if 'John Gabriel Borkman' in str(letter_text):
            all_corresp_ids[corresp_id].append('John Gabriel Borkman')
        if 'Kejser og Galilæer' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kejser og Galilæer')
        if 'Kjæmpehøien' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kjæmpehøien')
        if 'Kjærlighedens Komedie' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kjærlighedens Komedie')
        if 'Kongs-Emnerne' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kongs-Emnerne')
        if 'Lille Eyolf' in str(letter_text):
            all_corresp_ids[corresp_id].append('Lille Eyolf')
        if 'Norma.' in str(letter_text) or 'Norma ' in str(letter_text) or 'Norma»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Norma')
        if 'Når vi døde vågner' in str(letter_text):
            all_corresp_ids[corresp_id].append('Når vi døde vågner')
        if 'Olaf Liljekrans' in str(letter_text):
            all_corresp_ids[corresp_id].append('Olaf Liljekrans')
        if 'Peer Gynt' in str(letter_text):
            all_corresp_ids[corresp_id].append('Peer Gynt')
        if 'Rosmersholm' in str(letter_text):
            all_corresp_ids[corresp_id].append('Rosmersholm')
        if 'Rypen i Justedal' in str(letter_text):
            all_corresp_ids[corresp_id].append('Rypen i Justedal')
        if 'Samfundets støtter' in str(letter_text):
            all_corresp_ids[corresp_id].append('Samfundets støtter')
        if 'Sancthansnatten' in str(letter_text):
            all_corresp_ids[corresp_id].append('Sancthansnatten')
        if 'Svanhild.' in str(letter_text) or 'Svanhild ' in str(letter_text) or 'Svanhild»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Svanhild')
        if 'Vildanden' in str(letter_text):
            all_corresp_ids[corresp_id].append('Vildanden')
        if 'Terje Vigen' in str(letter_text):
            all_corresp_ids[corresp_id].append('Terje Vigen')
        if 'Digte.' in str(letter_text) or 'Digte ' in str(letter_text) or 'Digte»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Digte')
        else:
            continue

    for corresp in letter_text1871_1879:
        corresp_id = corresp.attrib['corresp']
        all_corresp_ids[corresp_id]=[]
        letter_text = corresp.xpath('.//tei:div[@type="letter"]/tei:p/text()', namespaces=ns)
        # print(letter_text)
        mentioned_titles = []
        if 'Brand.' in str(letter_text) or 'Brand ' in str(letter_text) or 'Brand»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Brand')
        if 'Bygmester Solness' in str(letter_text):
            all_corresp_ids[corresp_id].append('Bygmester Solness')
        if 'Catilina' in str(letter_text):
            all_corresp_ids[corresp_id].append('Catilina')
        if 'De unges Forbund' in str(letter_text):
            all_corresp_ids[corresp_id].append('De unges Forbund')
        if 'En folkefiende' in str(letter_text):
            all_corresp_ids[corresp_id].append('En folkefiende')
        if 'Episk Brand' in str(letter_text):
            all_corresp_ids[corresp_id].append('Episk Brand')
        if 'Et dukkehjem' in str(letter_text):
            all_corresp_ids[corresp_id].append('Et dukkehjem')
        if 'Fjeldfuglen' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fjeldfuglen')
        if 'Fruen fra havet' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fruen fra havet')
        if 'Fru Inger til Østeraad' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fru Inger til Østeraad')
        if 'Fru Inger til Østråt' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fru Inger til Østråt')
        if 'Gengangere' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gengangere')
        if 'Gildet paa Solhoug' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gildet paa Solhoug')
        if 'Gildet på Solhaug' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gildet på Solhaug')
        if 'Hedda Gabler' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hedda Gabler')
        if 'Hvide heste' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hvide heste')
        if 'Hærmændene paa Helgeland' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hærmændene paa Helgeland')
        if 'Hærmændene på Helgeland' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hærmændene på Helgeland')
        if 'John Gabriel Borkman' in str(letter_text):
            all_corresp_ids[corresp_id].append('John Gabriel Borkman')
        if 'Kejser og Galilæer' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kejser og Galilæer')
        if 'Kjæmpehøien' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kjæmpehøien')
        if 'Kjærlighedens Komedie' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kjærlighedens Komedie')
        if 'Kongs-Emnerne' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kongs-Emnerne')
        if 'Lille Eyolf' in str(letter_text):
            all_corresp_ids[corresp_id].append('Lille Eyolf')
        if 'Norma.' in str(letter_text) or 'Norma ' in str(letter_text) or 'Norma»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Norma')
        if 'Når vi døde vågner' in str(letter_text):
            all_corresp_ids[corresp_id].append('Når vi døde vågner')
        if 'Olaf Liljekrans' in str(letter_text):
            all_corresp_ids[corresp_id].append('Olaf Liljekrans')
        if 'Peer Gynt' in str(letter_text):
            all_corresp_ids[corresp_id].append('Peer Gynt')
        if 'Rosmersholm' in str(letter_text):
            all_corresp_ids[corresp_id].append('Rosmersholm')
        if 'Rypen i Justedal' in str(letter_text):
            all_corresp_ids[corresp_id].append('Rypen i Justedal')
        if 'Samfundets støtter' in str(letter_text):
            all_corresp_ids[corresp_id].append('Samfundets støtter')
        if 'Sancthansnatten' in str(letter_text):
            all_corresp_ids[corresp_id].append('Sancthansnatten')
        if 'Svanhild.' in str(letter_text) or 'Svanhild ' in str(letter_text) or 'Svanhild»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Svanhild')
        if 'Vildanden' in str(letter_text):
            all_corresp_ids[corresp_id].append('Vildanden')
        if 'Terje Vigen' in str(letter_text):
            all_corresp_ids[corresp_id].append('Terje Vigen')
        if 'Digte.' in str(letter_text) or 'Digte ' in str(letter_text) or 'Digte»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Digte')
        else:
            continue

    for corresp in letter_text1880_1889:
        corresp_id = corresp.attrib['corresp']
        all_corresp_ids[corresp_id]=[]
        letter_text = corresp.xpath('.//tei:div[@type="letter"]/tei:p/text()', namespaces=ns)
        # print(letter_text)
        mentioned_titles = []
        if 'Brand.' in str(letter_text) or 'Brand ' in str(letter_text) or 'Brand»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Brand')
        if 'Bygmester Solness' in str(letter_text):
            all_corresp_ids[corresp_id].append('Bygmester Solness')
        if 'Catilina' in str(letter_text):
            all_corresp_ids[corresp_id].append('Catilina')
        if 'De unges Forbund' in str(letter_text):
            all_corresp_ids[corresp_id].append('De unges Forbund')
        if 'En folkefiende' in str(letter_text):
            all_corresp_ids[corresp_id].append('En folkefiende')
        if 'Episk Brand' in str(letter_text):
            all_corresp_ids[corresp_id].append('Episk Brand')
        if 'Et dukkehjem' in str(letter_text):
            all_corresp_ids[corresp_id].append('Et dukkehjem')
        if 'Fjeldfuglen' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fjeldfuglen')
        if 'Fruen fra havet' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fruen fra havet')
        if 'Fru Inger til Østeraad' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fru Inger til Østeraad')
        if 'Fru Inger til Østråt' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fru Inger til Østråt')
        if 'Gengangere' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gengangere')
        if 'Gildet paa Solhoug' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gildet paa Solhoug')
        if 'Gildet på Solhaug' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gildet på Solhaug')
        if 'Hedda Gabler' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hedda Gabler')
        if 'Hvide heste' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hvide heste')
        if 'Hærmændene paa Helgeland' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hærmændene paa Helgeland')
        if 'Hærmændene på Helgeland' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hærmændene på Helgeland')
        if 'John Gabriel Borkman' in str(letter_text):
            all_corresp_ids[corresp_id].append('John Gabriel Borkman')
        if 'Kejser og Galilæer' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kejser og Galilæer')
        if 'Kjæmpehøien' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kjæmpehøien')
        if 'Kjærlighedens Komedie' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kjærlighedens Komedie')
        if 'Kongs-Emnerne' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kongs-Emnerne')
        if 'Lille Eyolf' in str(letter_text):
            all_corresp_ids[corresp_id].append('Lille Eyolf')
        if 'Norma.' in str(letter_text) or 'Norma ' in str(letter_text) or 'Norma»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Norma')
        if 'Når vi døde vågner' in str(letter_text):
            all_corresp_ids[corresp_id].append('Når vi døde vågner')
        if 'Olaf Liljekrans' in str(letter_text):
            all_corresp_ids[corresp_id].append('Olaf Liljekrans')
        if 'Peer Gynt' in str(letter_text):
            all_corresp_ids[corresp_id].append('Peer Gynt')
        if 'Rosmersholm' in str(letter_text):
            all_corresp_ids[corresp_id].append('Rosmersholm')
        if 'Rypen i Justedal' in str(letter_text):
            all_corresp_ids[corresp_id].append('Rypen i Justedal')
        if 'Samfundets støtter' in str(letter_text):
            all_corresp_ids[corresp_id].append('Samfundets støtter')
        if 'Sancthansnatten' in str(letter_text):
            all_corresp_ids[corresp_id].append('Sancthansnatten')
        if 'Svanhild.' in str(letter_text) or 'Svanhild ' in str(letter_text) or 'Svanhild»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Svanhild')
        if 'Vildanden' in str(letter_text):
            all_corresp_ids[corresp_id].append('Vildanden')
        if 'Terje Vigen' in str(letter_text):
            all_corresp_ids[corresp_id].append('Terje Vigen')
        if 'Digte.' in str(letter_text) or 'Digte ' in str(letter_text) or 'Digte»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Digte')
        else:
            continue

    for corresp in letter_text1890_1905:
        corresp_id = corresp.attrib['corresp']
        all_corresp_ids[corresp_id]=[]
        letter_text = corresp.xpath('.//tei:div[@type="letter"]/tei:p/text()', namespaces=ns)
        # print(letter_text)
        mentioned_titles = []
        if 'Brand.' in str(letter_text) or 'Brand ' in str(letter_text) or 'Brand»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Brand')
        if 'Bygmester Solness' in str(letter_text):
            all_corresp_ids[corresp_id].append('Bygmester Solness')
        if 'Catilina' in str(letter_text):
            all_corresp_ids[corresp_id].append('Catilina')
        if 'De unges Forbund' in str(letter_text):
            all_corresp_ids[corresp_id].append('De unges Forbund')
        if 'En folkefiende' in str(letter_text):
            all_corresp_ids[corresp_id].append('En folkefiende')
        if 'Episk Brand' in str(letter_text):
            all_corresp_ids[corresp_id].append('Episk Brand')
        if 'Et dukkehjem' in str(letter_text):
            all_corresp_ids[corresp_id].append('Et dukkehjem')
        if 'Fjeldfuglen' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fjeldfuglen')
        if 'Fruen fra havet' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fruen fra havet')
        if 'Fru Inger til Østeraad' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fru Inger til Østeraad')
        if 'Fru Inger til Østråt' in str(letter_text):
            all_corresp_ids[corresp_id].append('Fru Inger til Østråt')
        if 'Gengangere' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gengangere')
        if 'Gildet paa Solhoug' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gildet paa Solhoug')
        if 'Gildet på Solhaug' in str(letter_text):
            all_corresp_ids[corresp_id].append('Gildet på Solhaug')
        if 'Hedda Gabler' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hedda Gabler')
        if 'Hvide heste' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hvide heste')
        if 'Hærmændene paa Helgeland' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hærmændene paa Helgeland')
        if 'Hærmændene på Helgeland' in str(letter_text):
            all_corresp_ids[corresp_id].append('Hærmændene på Helgeland')
        if 'John Gabriel Borkman' in str(letter_text):
            all_corresp_ids[corresp_id].append('John Gabriel Borkman')
        if 'Kejser og Galilæer' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kejser og Galilæer')
        if 'Kjæmpehøien' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kjæmpehøien')
        if 'Kjærlighedens Komedie' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kjærlighedens Komedie')
        if 'Kongs-Emnerne' in str(letter_text):
            all_corresp_ids[corresp_id].append('Kongs-Emnerne')
        if 'Lille Eyolf' in str(letter_text):
            all_corresp_ids[corresp_id].append('Lille Eyolf')
        if 'Norma.' in str(letter_text) or 'Norma ' in str(letter_text) or 'Norma»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Norma')
        if 'Når vi døde vågner' in str(letter_text):
            all_corresp_ids[corresp_id].append('Når vi døde vågner')
        if 'Olaf Liljekrans' in str(letter_text):
            all_corresp_ids[corresp_id].append('Olaf Liljekrans')
        if 'Peer Gynt' in str(letter_text):
            all_corresp_ids[corresp_id].append('Peer Gynt')
        if 'Rosmersholm' in str(letter_text):
            all_corresp_ids[corresp_id].append('Rosmersholm')
        if 'Rypen i Justedal' in str(letter_text):
            all_corresp_ids[corresp_id].append('Rypen i Justedal')
        if 'Samfundets støtter' in str(letter_text):
            all_corresp_ids[corresp_id].append('Samfundets støtter')
        if 'Sancthansnatten' in str(letter_text):
            all_corresp_ids[corresp_id].append('Sancthansnatten')
        if 'Svanhild.' in str(letter_text) or 'Svanhild ' in str(letter_text) or 'Svanhild»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Svanhild')
        if 'Vildanden' in str(letter_text):
            all_corresp_ids[corresp_id].append('Vildanden')
        if 'Terje Vigen' in str(letter_text):
            all_corresp_ids[corresp_id].append('Terje Vigen')
        if 'Digte.' in str(letter_text) or 'Digte ' in str(letter_text) or 'Digte»' in str(letter_text):
            all_corresp_ids[corresp_id].append('Digte')
        else:
            continue

    # pprint.pprint(all_corresp_ids)
    return all_corresp_ids


# ids_and_title_from_letters(letter_text1844_1871)
ids_titles_dict = ids_and_title_from_letters(letter_text1844_1871)


# compare the data collected above with that contained in the csv. Add missing information
def edit_works_info(dict_titles_in_letters, letter_id_csv, title_in_csv, genres_csv, title_ids, collective_title_csv, works_ids_csv):
    workbook = openpyxl.load_workbook('../xml-filer/Works_ids_and_text_NME.xlsx')
    sheet = workbook.active
    verkID = []
    tittel = []
    tekstID = []
    sjanger = []

    for rowNum in range(2, 70):
        verkID.append(sheet.cell(row=rowNum, column=1).value)
        tittel.append(re.sub('«', '', re.sub('»','',sheet.cell(row=rowNum, column=3).value)))
        tekstID.append(sheet.cell(row=rowNum, column=4).value)
        sjanger.append(sheet.cell(row=rowNum, column=5).value)
    # print(len(letter_id_csv))
    # print(len(title_in_csv))
    # print(len(genres_csv))
    # print(len(title_ids))
    # print(len(collective_title_csv))
    new_work_title = []
    new_genres = []
    new_title_ids = []
    new_collective_titles = []
    new_work_ids = []

    for letter_id, title_csv, genre, title_id, collective_title, work_id in zip(letter_id_csv, title_in_csv, genres_csv, title_ids, collective_title_csv, works_ids_csv):

        try:
            if len(dict_titles_in_letters[letter_id])!=0:
                add_titles = re.sub('\[', '', re.sub('\]', '', re.sub('\'', '', str(dict_titles_in_letters[letter_id]))))
                title_tokens = add_titles.split(', ')
                for token in title_tokens:
                    if token in title_csv:
                        continue
                    else:
                        if title_csv == '':
                            try:
                                index = tittel.index(token)
                                title_csv = (str(token))
                                genre = str(sjanger[index])
                                title_id = str(tekstID[index])
                                work_id = str(verkID[index])
                                if any(c in str(tekstID[index]) for c in ('C1', 'C2')):
                                    collective_title = str('C')
                                elif any(c in str(tekstID[index]) for c in ('F1', 'F2')):
                                    collective_title = str('F')
                                elif any(c in str(tekstID[index]) for c in ('H1', 'H2')):
                                    collective_title = str('HH')
                                elif any(c in str(tekstID[index]) for c in ('G1', 'G2')):
                                    collective_title = str('G')
                                elif any(c in str(tekstID[index]) for c in ('K1', 'K2')):
                                    collective_title = str('K')
                                elif any(c in str(tekstID[index]) for c in ('OL1', 'OL2', 'OL3')):
                                    collective_title = str('OL')
                                else:
                                    collective_title = str('')
                            except ValueError:
                                title_csv = str(token)
                                genre = str('drama')
                                title_id = str('')
                                collective_title = str('')
                                work_id = str('')
                        else:
                            try:
                                index = tittel.index(token)
                                title_csv = str(title_csv + str(', ') + token)
                                genre = str(genre +str(', ')+str(sjanger[index]))
                                title_id = str(title_id + str(', ')+str(tekstID[index]))
                                work_id = str(work_id + str(', ')+str(verkID[index]))
                                if len(collective_title)==0:
                                    if any(c in str(tekstID[index]) for c in ('C1', 'C2')):
                                        collective_title = str('C')
                                    elif any(c in str(tekstID[index]) for c in ('F1', 'F2')):
                                        collective_title =str('F')
                                    elif any(c in str(tekstID[index]) for c in ('H1', 'H2')):
                                        collective_title = str('HH')
                                    elif any(c in str(tekstID[index]) for c in ('G1', 'G2')):
                                        collective_title =str('G')
                                    elif any(c in str(tekstID[index]) for c in ('K1', 'K2')):
                                        collective_title =str('K')
                                    elif any(c in str(tekstID[index]) for c in ('OL1', 'OL2', 'OL3')):
                                        collective_title = str('OL')
                                    # else:
                                    #     collective_title = str('')
                                else:
                                    if any(c in str(tekstID[index]) for c in ('C1', 'C2')):
                                        collective_title = str(collective_title + str(', ') + str('C'))
                                    elif any(c in str(tekstID[index]) for c in ('F1', 'F2')):
                                        collective_title = str(collective_title + str(', ') + str('F'))
                                    elif any(c in str(tekstID[index]) for c in ('H1', 'H2')):
                                        collective_title = str(collective_title + str(', ') + str('HH'))
                                    elif any(c in str(tekstID[index]) for c in ('G1', 'G2')):
                                        collective_title = str(collective_title + str(', ') + str('G'))
                                    elif any(c in str(tekstID[index]) for c in ('K1', 'K2')):
                                        collective_title = str(collective_title + str(', ') + str('K'))
                                    elif any(c in str(tekstID[index]) for c in ('OL1', 'OL2', 'OL3')):
                                        collective_title = str(collective_title + str(', ') + str('OL'))
                                    # else:
                                    #     collective_title = str('')
                            except ValueError:
                                title_csv = str(title_csv + str(', ') + token)
                                genre = str(genre + str(', drama'))
                                title_id = str(title_id + str(', '))
                                collective_title = str(collective_title + str(''))
                                work_id = str(work_id + str(''))
                #print(add_titles)
                new_work_title.append(title_csv)
                new_genres.append(genre)
                new_title_ids.append(title_id)
                new_collective_titles.append(collective_title)
                new_work_ids.append(work_id)

            else:
                new_work_title.append(title_csv)
                new_genres.append(genre)
                new_title_ids.append(title_id)
                new_collective_titles.append(collective_title)
                new_work_ids.append(work_id)

        except KeyError:
            new_work_title.append(title_csv)
            new_genres.append(genre)
            new_title_ids.append(title_id)
            new_collective_titles.append(collective_title)
            new_work_ids.append(work_id)

    pprint.pprint(new_work_ids)
    print(len(new_work_ids))
    return new_work_title, new_genres, new_title_ids, new_collective_titles, new_work_ids


new_work_title, new_genres, new_title_ids, new_collective_titles, new_works_ids = edit_works_info(ids_titles_dict,Letter_ID,Mention_Works_Title,Mention_Works_Genre,Text_ID,Collective_Vesions_ID, Mentioned_Works_IDs)
# edit_works_info(ids_titles_dict,Letter_ID,Mention_Works_Title,Mention_Works_Genre,Text_ID,
# Collective_Vesions_ID, Mentioned_Works_IDs)


# compile all data -newly edited and none edited- and create new csv file
def create_csv(new_work_title, new_genres, new_title_ids, new_collective_titles, new_works_ids):
    colnames = ['Letter_ID', 'Sender_ID', 'Sender_Name', 'Recipient_ID', 'Recipient_Name', 'Date',
                'Dispatch_Location', 'Dispatch_Location_Abbr', 'GeoName_ID', 'toponymName', 'Country',
                'Latitude', 'Longitude', 'Recipient_Adress', 'Recipient_Location', 'Recipient_Location_Abbr',
                'Recipient_Addr_Latitude', 'Recipient_Addr_Longitude', 'Mentioned_Person_ID',
                'Mentioned_Persons', 'Mentioned_Org_ID', 'Mentioned_Org', 'Mentioned_Works_IDs',
                'Mention_Works_Title', 'Text_ID', 'Mention_Works_Genre', 'Collective_Vesions_ID',
                'Mentioned_Places', 'Mentioned_Places_Abbr']
    work_file = pd.read_csv('Compiled_Letter_Data.csv', names=colnames, na_filter=False)

    Letter_ID = work_file.Letter_ID.tolist()
    Sender_ID = work_file.Sender_ID.tolist()
    Sender_Name = work_file.Sender_Name.tolist()
    Recipient_ID = work_file.Recipient_ID.tolist()
    Recipient_Name = work_file.Recipient_Name.tolist()
    Date = work_file.Date.tolist()
    Dispatch_Location = work_file.Dispatch_Location.tolist()
    Dispatch_Location_Abbr = work_file.Dispatch_Location_Abbr.tolist()
    GeoName_ID = work_file.GeoName_ID.tolist()
    toponymName = work_file.toponymName.tolist()
    Country = work_file.Country.tolist()
    Latitude = work_file.Latitude.tolist()
    Longitude = work_file.Longitude.tolist()
    Recipient_Adress = work_file.Recipient_Adress.tolist()
    Recipient_Location = work_file.Recipient_Location.tolist()
    Recipient_Location_Abbr = work_file.Recipient_Location_Abbr.tolist()
    Recipient_Addr_Latitude = work_file.Recipient_Addr_Latitude.tolist()
    Recipient_Addr_Longitude = work_file.Recipient_Addr_Longitude.tolist()
    Mentioned_Person_ID = work_file.Mentioned_Person_ID.tolist()
    Mentioned_Persons = work_file.Mentioned_Persons.tolist()
    Mentioned_Org_ID = work_file.Mentioned_Org_ID.tolist()
    Mentioned_Org = work_file.Mentioned_Org.tolist()
    Mentioned_Places = work_file.Mentioned_Places.tolist()
    Mentioned_Places_Abbr = work_file.Mentioned_Places_Abbr.tolist()


    rows = zip(Letter_ID,Sender_ID,Sender_Name,Recipient_ID,Recipient_Name,Date,
            Dispatch_Location,Dispatch_Location_Abbr,GeoName_ID,toponymName,Country,
            Latitude,Longitude,Recipient_Adress,Recipient_Location,Recipient_Location_Abbr,
            Recipient_Addr_Latitude,Recipient_Addr_Longitude,Mentioned_Person_ID,
            Mentioned_Persons,Mentioned_Org_ID,Mentioned_Org,new_works_ids,
            new_work_title,new_title_ids,new_genres,new_collective_titles,
            Mentioned_Places,Mentioned_Places_Abbr)

    with open('Compiled_Letter_Data.csv', 'w', ) as work_csv:
        wr = csv.writer(work_csv, delimiter=',')

        for row in rows:
            wr.writerow(row)


create_csv(new_work_title, new_genres, new_title_ids, new_collective_titles, new_works_ids)

